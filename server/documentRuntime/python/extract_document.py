from __future__ import annotations

import csv
import importlib
import importlib.metadata
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple


LIST_MARKER_PATTERN = re.compile(r"^(\*|-|\u2022|\d+[\.\)])\s+")
HEADING_PREFIX_PATTERN = re.compile(r"^(\d+(?:\.\d+)*\.?)\s+")
MARKDOWN_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.*)$")
MARKDOWN_TABLE_SEPARATOR_PATTERN = re.compile(r"^\|\s*:?-{3,}")


def load_json(path: str) -> Dict[str, Any]:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def write_json(path: str, value: Dict[str, Any]) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def module_available(name: str) -> bool:
    return importlib.util.find_spec(name) is not None  # type: ignore[attr-defined]


def module_version(name: str) -> Optional[str]:
    try:
        return importlib.metadata.version(name)
    except Exception:
        return None


def build_capabilities() -> Dict[str, Any]:
    modules = {
        "pypdf": module_available("pypdf"),
        "docx": module_available("docx"),
        "openpyxl": module_available("openpyxl"),
        "paddleocr": module_available("paddleocr"),
        "docling": module_available("docling"),
    }
    versions = {
        "pypdf": module_version("pypdf"),
        "python-docx": module_version("python-docx"),
        "openpyxl": module_version("openpyxl"),
        "paddleocr": module_version("paddleocr"),
        "docling": module_version("docling"),
    }
    reasons: List[str] = []
    if not modules["pypdf"]:
        reasons.append("pypdf is unavailable.")
    if not modules["docx"]:
        reasons.append("python-docx is unavailable.")
    if not modules["openpyxl"]:
        reasons.append("openpyxl is unavailable.")
    if not modules["paddleocr"]:
        reasons.append("PaddleOCR is unavailable.")
    if not modules["docling"]:
        reasons.append("Docling is unavailable.")
    return {
        "available": any(modules.values()),
        "nativeReady": modules["pypdf"] or modules["docx"] or modules["openpyxl"],
        "ocrReady": modules["paddleocr"] or modules["docling"],
        "modules": modules,
        "versions": versions,
        "reasons": reasons,
    }


def normalize_text(value: str, *, preserve_line_breaks: bool = True) -> str:
    text = str(value or "").replace("\r", "").replace("\u00a0", " ")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    if preserve_line_breaks:
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()
    return re.sub(r"\s+", " ", text).strip()


def normalize_pdf_text(value: str) -> str:
    lines: List[str] = []
    for raw_line in str(value or "").replace("\r", "").split("\n"):
        line = re.sub(r"[ \t]{2,}", " ", raw_line.replace("\u00a0", " ")).strip()
        if not line:
            if lines and lines[-1] != "":
                lines.append("")
            continue
        if lines and lines[-1].endswith("-") and line[:1].islower():
            lines[-1] = f"{lines[-1][:-1]}{line}"
            continue
        lines.append(line)
    return normalize_text("\n".join(lines))


def build_language_hints(text: str) -> List[str]:
    hints: List[str] = []
    if re.search(r"[A-Za-z]", text or ""):
        hints.append("en")
    if re.search(r"[\u0600-\u06FF]", text or ""):
        hints.append("ar")
    return hints


def normalize_table_rows(rows: Iterable[Sequence[Any]]) -> List[List[str]]:
    normalized_rows: List[List[str]] = []
    max_len = 0
    for row in rows:
        converted = [normalize_text(str(cell), preserve_line_breaks=False) if cell is not None else "" for cell in row]
        while converted and converted[-1] == "":
            converted.pop()
        if not any(converted):
            continue
        max_len = max(max_len, len(converted))
        normalized_rows.append(converted)
    for row in normalized_rows:
        if len(row) < max_len:
            row.extend([""] * (max_len - len(row)))
    return normalized_rows


def looks_like_heading(line: str) -> bool:
    trimmed = line.strip()
    if not trimmed or len(trimmed) > 120:
        return False
    if re.match(r"^(chapter|section|unit|module|lesson|topic|abstract|introduction|conclusion)\b", trimmed, re.I):
        return True
    if HEADING_PREFIX_PATTERN.match(trimmed):
        return True
    if re.match(r"^[A-Z0-9\s\-:()]{4,}$", trimmed):
        return True
    if re.match(r"^[\u0600-\u06FF0-9\s\-:()]{4,}$", trimmed) and len(trimmed) <= 80:
        return True
    return False


def looks_like_title_candidate(line: str, index: int) -> bool:
    trimmed = line.strip()
    if index > 1 or not trimmed or len(trimmed) > 100:
        return False
    return bool(
        re.match(r"^[A-Z][A-Za-z0-9\s\-:()]{4,}$", trimmed)
        or re.match(r"^[\u0600-\u06FF][\u0600-\u06FF0-9\s\-:()]{3,}$", trimmed)
    )


def looks_like_list_item(line: str) -> bool:
    return bool(LIST_MARKER_PATTERN.match(line.strip()))


def looks_like_table_row(line: str) -> bool:
    trimmed = line.strip()
    return bool(trimmed) and ("\t" in trimmed or "|" in trimmed or re.search(r"\s{3,}", trimmed) is not None)


def infer_heading_metadata(line: str, index: int) -> Optional[Tuple[str, int]]:
    trimmed = line.strip()
    if not looks_like_heading(trimmed):
        if looks_like_title_candidate(trimmed, index):
            return ("title", 1)
        return None

    numbered_match = HEADING_PREFIX_PATTERN.match(trimmed)
    if numbered_match:
        depth = len(numbered_match.group(1).rstrip(".").split("."))
        return ("heading" if depth <= 1 else "subheading", min(6, depth + 1))

    if looks_like_title_candidate(trimmed, index):
        return ("title", 1)

    return ("heading" if re.match(r"^[A-Z0-9\s\-:()]{4,}$", trimmed) else "subheading", 2)


def create_block(
    page_number: int,
    order: int,
    text: str,
    source: str,
    *,
    block_type: Optional[str] = None,
    level: Optional[int] = None,
    rows: Optional[List[List[str]]] = None,
    confidence: Optional[float] = None,
    bbox: Optional[Dict[str, float]] = None,
    notes: Optional[List[str]] = None,
) -> Dict[str, Any]:
    normalized = normalize_text(text)
    resolved_type = block_type or "paragraph"
    resolved_level = level
    resolved_rows = rows
    if not block_type:
        heading = infer_heading_metadata(normalized, order - 1)
        if heading:
            resolved_type, resolved_level = heading
        elif looks_like_list_item(normalized):
            resolved_type = "list_item"
        elif looks_like_table_row(normalized):
            resolved_type = "table"
            resolved_rows = normalize_table_rows([re.split(r"\s{3,}|\t|\|", normalized)])
    return {
        "blockId": f"py-{page_number}-{order}",
        "type": resolved_type,
        "source": source,
        "text": normalized,
        "pageNumber": page_number,
        "order": order,
        "level": resolved_level,
        "rows": resolved_rows,
        "confidence": confidence,
        "bbox": bbox,
        "notes": notes,
    }


def block_plain_text(block: Dict[str, Any]) -> str:
    block_type = str(block.get("type") or "paragraph")
    text = str(block.get("text") or "").strip()
    if not text:
        return ""
    if block_type == "list_item":
        return f"- {LIST_MARKER_PATTERN.sub('', text).strip()}"
    if block_type == "table":
        rows = block.get("rows") or []
        if isinstance(rows, list) and rows:
            return "\n".join(" | ".join(cell for cell in row if cell) for row in rows)
    if block_type == "ocr_block":
        return f"OCR: {text}"
    if block_type == "note":
        return f"Note: {text}"
    return text


def lines_to_blocks(page_number: int, text: str, source: str, starting_order: int = 0) -> List[Dict[str, Any]]:
    lines = [line.strip() for line in normalize_text(text).split("\n") if line.strip()]
    return [create_block(page_number, starting_order + index, line, source) for index, line in enumerate(lines, start=1)]


def page_from_blocks(
    page_number: int,
    blocks: List[Dict[str, Any]],
    source_kind: str,
    *,
    unit_label: Optional[str] = None,
    unit_kind: str = "page",
) -> Dict[str, Any]:
    sorted_blocks = sorted(blocks, key=lambda item: int(item.get("order") or 0))
    page_text = normalize_text("\n".join(block_plain_text(block) for block in sorted_blocks if block_plain_text(block)))
    headings = [
        str(block.get("text") or "")
        for block in sorted_blocks
        if str(block.get("type") or "") in ("title", "heading", "subheading")
    ][:10]
    return {
        "pageNumber": page_number,
        "text": page_text,
        "headingCandidates": headings,
        "blocks": sorted_blocks,
        "sourceKind": source_kind,
        "unitLabel": unit_label or f"Page {page_number}",
        "unitKind": unit_kind,
        "charCount": len(page_text),
    }


def build_result_payload(
    *,
    engine: str,
    pages: List[Dict[str, Any]],
    full_text: str,
    notes: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
    extractor_chain: Optional[List[str]] = None,
    fallback_chain: Optional[List[str]] = None,
    ocr_blocks: Optional[List[Dict[str, Any]]] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    normalized_full_text = normalize_text(full_text)
    payload = {
        "engine": engine,
        "pages": pages,
        "fullText": normalized_full_text,
        "languageHints": build_language_hints(normalized_full_text),
        "notes": notes or [],
        "warnings": warnings or [],
        "extractorChain": extractor_chain or [engine],
        "fallbackChain": fallback_chain or [],
        "pageCount": len(pages),
        "textLength": len(normalized_full_text),
        "tablesDetected": sum(1 for page in pages for block in page.get("blocks") or [] if block.get("type") == "table"),
        "listsDetected": sum(1 for page in pages for block in page.get("blocks") or [] if block.get("type") == "list_item"),
        "pageSummaries": [
            {
                "pageNumber": page.get("pageNumber"),
                "unitLabel": page.get("unitLabel"),
                "unitKind": page.get("unitKind"),
                "charCount": page.get("charCount"),
                "blockCount": len(page.get("blocks") or []),
            }
            for page in pages
        ],
        "metadata": metadata or {},
    }
    if ocr_blocks is not None:
        payload["ocrBlocks"] = ocr_blocks
    return payload


def extract_pdf_native(source_path: str) -> Dict[str, Any]:
    from pypdf import PdfReader  # type: ignore

    reader = PdfReader(source_path)
    pages: List[Dict[str, Any]] = []
    notes: List[str] = []
    weak_pages: List[int] = []
    for index, page in enumerate(reader.pages, start=1):
        text = ""
        try:
            text = page.extract_text(extraction_mode="layout") or ""
            if text.strip():
                notes.append(f"Page {index}: pypdf layout extraction succeeded.")
        except TypeError:
            pass
        except Exception as error:
            notes.append(f"Page {index}: pypdf layout extraction failed: {error}")
        if not text.strip():
            try:
                text = page.extract_text() or ""
                if text.strip():
                    notes.append(f"Page {index}: pypdf plain extraction fallback succeeded.")
            except Exception as error:
                notes.append(f"Page {index}: pypdf plain extraction failed: {error}")
        page_payload = page_from_blocks(index, lines_to_blocks(index, normalize_pdf_text(text), "native"), "native")
        if len(str(page_payload.get("text") or "")) < PDF_WEAK_PAGE_CHAR_THRESHOLD:
            weak_pages.append(index)
        pages.append(page_payload)
    full_text = normalize_text("\n\n".join(page["text"] for page in pages if page["text"]))
    warnings = [
        "Low-text PDF pages detected and should be considered for OCR fallback: "
        + ", ".join(str(page_number) for page_number in weak_pages)
    ] if weak_pages else []
    return build_result_payload(
        engine="python:pypdf",
        pages=pages,
        full_text=full_text,
        notes=notes,
        warnings=warnings,
        extractor_chain=["python:pypdf:layout", "python:pypdf:plain-fallback"],
        metadata={"weakPageNumbers": weak_pages},
    )


def extract_docx_native(source_path: str) -> Dict[str, Any]:
    from docx import Document  # type: ignore
    from docx.document import Document as DocxDocument  # type: ignore
    from docx.oxml.table import CT_Tbl  # type: ignore
    from docx.oxml.text.paragraph import CT_P  # type: ignore
    from docx.table import Table  # type: ignore
    from docx.text.paragraph import Paragraph  # type: ignore

    document = Document(source_path)
    blocks: List[Dict[str, Any]] = []
    order = 1
    body = document.element.body if isinstance(document, DocxDocument) else document._element  # type: ignore[attr-defined]
    for child in body.iterchildren():
        if isinstance(child, CT_P):
            paragraph = Paragraph(child, document)
            text = normalize_text(paragraph.text or "")
            if not text:
                continue
            style_name = getattr(getattr(paragraph, "style", None), "name", "") or ""
            style = style_name.strip().lower()
            if style.startswith("title"):
                block = create_block(1, order, text, "native", block_type="title", level=1)
            elif style.startswith("heading"):
                match = re.search(r"(\d+)", style)
                level = int(match.group(1)) if match else 1
                block = create_block(1, order, text, "native", block_type="heading" if level <= 1 else "subheading", level=min(6, level + 1))
            elif "list" in style or getattr(getattr(paragraph, "_p", None), "pPr", None) is not None and getattr(getattr(getattr(paragraph, "_p", None), "pPr", None), "numPr", None) is not None:
                block = create_block(1, order, text, "native", block_type="list_item")
            else:
                block = create_block(1, order, text, "native")
            blocks.append(block)
            order += 1
        elif isinstance(child, CT_Tbl):
            table = Table(child, document)
            rows = normalize_table_rows([[cell.text for cell in row.cells] for row in table.rows])
            if not rows:
                continue
            blocks.append(create_block(1, order, "\n".join(" | ".join(cell for cell in row if cell) for row in rows), "native", block_type="table", rows=rows, notes=["DOCX table preserved in document body order."]))
            order += 1
    pages = [page_from_blocks(1, blocks, "native", unit_label="Section 1: Document Body", unit_kind="section")]
    return build_result_payload(
        engine="python:python-docx",
        pages=pages,
        full_text=pages[0]["text"] if pages else "",
        warnings=["DOCX document body produced no readable blocks."] if not blocks else [],
        extractor_chain=["python:python-docx:body-order"],
    )


def build_sheet_headers(row: Sequence[str]) -> List[str]:
    seen: Dict[str, int] = {}
    headers: List[str] = []
    for index, cell in enumerate(row, start=1):
        base = normalize_text(cell, preserve_line_breaks=False) or f"Column {index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    return headers


def detect_sheet_header_row(rows: Sequence[Sequence[str]]) -> Optional[int]:
    if not rows or len(rows[0]) < 2:
        return None
    first_row = [cell for cell in rows[0] if cell]
    if len(first_row) < 2:
        return None
    return 0 if len(set(first_row)) / max(1, len(first_row)) >= 0.6 else None


def summarize_sheet_row(headers: Sequence[str], row: Sequence[str], row_number: int) -> str:
    parts = [f"{headers[index]} = {value}" for index, value in enumerate(row) if index < len(headers) and value]
    return f"Row {row_number}: " + "; ".join(parts) if parts else ""


def extract_spreadsheet_native(source_path: str, file_type: str) -> Dict[str, Any]:
    pages: List[Dict[str, Any]] = []
    if file_type == "csv":
        with open(source_path, "r", encoding="utf-8", errors="ignore", newline="") as handle:
            sheet_rows = [("CSV Import", normalize_table_rows(csv.reader(handle)))]
    else:
        from openpyxl import load_workbook  # type: ignore
        workbook = load_workbook(source_path, data_only=True, read_only=True)
        sheet_rows = [(sheet_name, normalize_table_rows(workbook[sheet_name].iter_rows(values_only=True))) for sheet_name in workbook.sheetnames]

    for page_number, (sheet_name, rows) in enumerate(sheet_rows, start=1):
        blocks: List[Dict[str, Any]] = [
            create_block(page_number, 1, f"Sheet {page_number}: {sheet_name}", "native", block_type="heading", level=2),
            create_block(page_number, 2, f"Worksheet summary: {len(rows)} populated rows, {max((len(row) for row in rows), default=0)} columns.", "native", block_type="note", notes=["Spreadsheet rows stay in worksheet order for prompt reuse."]),
        ]
        order = 3
        header_row_index = detect_sheet_header_row(rows)
        headers = build_sheet_headers(rows[header_row_index]) if header_row_index is not None else []
        table_rows = rows if header_row_index is not None else ([build_sheet_headers(rows[0])] + rows if rows else [])
        for start in range(0, len(table_rows), 40):
            chunk = table_rows[start : start + 40]
            if chunk:
                blocks.append(create_block(page_number, order, "Worksheet grid", "native", block_type="table", rows=chunk))
                order += 1
        if header_row_index is not None and len(rows) <= 200:
            for relative_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
                summary = summarize_sheet_row(headers, row, relative_index)
                if summary:
                    blocks.append(create_block(page_number, order, summary, "native"))
                    order += 1
        elif header_row_index is None:
            for relative_index, row in enumerate(rows, start=1):
                summary = normalize_text(" | ".join(cell for cell in row if cell))
                if summary:
                    blocks.append(create_block(page_number, order, f"Row {relative_index}: {summary}", "native"))
                    order += 1
        pages.append(page_from_blocks(page_number, blocks, "native", unit_label=f"Sheet {page_number}: {sheet_name}", unit_kind="sheet"))

    full_text = normalize_text("\n\n".join(page["text"] for page in pages if page["text"]))
    return build_result_payload(engine="python:openpyxl" if file_type != "csv" else "python:csv", pages=pages, full_text=full_text, warnings=["Spreadsheet document contained no populated worksheets or rows."] if not pages else [], extractor_chain=["python:openpyxl:worksheet-native" if file_type != "csv" else "python:csv:native"], metadata={"unitKind": "sheet", "sheetCount": len(pages)})


def extract_plain_text(source_path: str) -> Dict[str, Any]:
    text = normalize_text(Path(source_path).read_text(encoding="utf-8", errors="ignore"))
    pages = [page_from_blocks(1, lines_to_blocks(1, text, "native"), "native")]
    return build_result_payload(engine="python:text", pages=pages, full_text=text, extractor_chain=["python:text:native"])


def build_paddleocr_engine() -> Tuple[Any, List[str]]:
    from paddleocr import PaddleOCR  # type: ignore
    for lang in ("ar", "arabic", "en"):
        try:
            return PaddleOCR(use_angle_cls=True, lang=lang), [f"PaddleOCR initialized with language profile: {lang}"]
        except Exception:
            continue
    return PaddleOCR(use_angle_cls=True, lang="en"), ["PaddleOCR fell back to the default English profile."]


def iter_ocr_pages(result: Any) -> List[List[Any]]:
    if not isinstance(result, list):
        return []
    if result and all(isinstance(entry, (list, tuple)) and len(entry) >= 2 for entry in result):
        return [list(result)]
    return [page for page in result if isinstance(page, list)]


def bbox_from_points(points: Sequence[Sequence[float]]) -> Optional[Dict[str, float]]:
    x_values = [float(point[0]) for point in points if len(point) >= 2]
    y_values = [float(point[1]) for point in points if len(point) >= 2]
    if not x_values or not y_values:
        return None
    return {"x": min(x_values), "y": min(y_values), "width": max(x_values) - min(x_values), "height": max(y_values) - min(y_values)}


def ocr_entry_to_block(page_number: int, order: int, entry: Sequence[Any]) -> Optional[Dict[str, Any]]:
    if len(entry) < 2:
        return None
    bbox = entry[0] if isinstance(entry[0], (list, tuple)) else []
    payload = entry[1] if isinstance(entry[1], (list, tuple)) else [entry[1], None]
    text = normalize_text(str(payload[0] if len(payload) > 0 else ""))
    if not text:
        return None
    try:
        confidence = float(payload[1]) if len(payload) > 1 and payload[1] is not None else None
    except Exception:
        confidence = None
    heading = infer_heading_metadata(text, order - 1)
    block_type = "ocr_block"
    level = None
    if heading:
        block_type, level = heading
    elif looks_like_list_item(text):
        block_type = "list_item"
    elif looks_like_table_row(text):
        block_type = "table"
    rows = normalize_table_rows([re.split(r"\s{3,}|\t|\|", text)]) if block_type == "table" else None
    return create_block(page_number, order, text, "ocr", block_type=block_type, level=level, rows=rows, confidence=confidence, bbox=bbox_from_points(bbox))


def extract_ocr_document(source_path: str, file_type: str) -> Dict[str, Any]:
    engine, notes = build_paddleocr_engine()
    result = engine.ocr(source_path, cls=True)
    pages: List[Dict[str, Any]] = []
    ocr_blocks: List[Dict[str, Any]] = []
    for page_number, entries in enumerate(iter_ocr_pages(result), start=1):
        page_blocks: List[Dict[str, Any]] = []
        for order, entry in enumerate(entries, start=1):
            block = ocr_entry_to_block(page_number, order, entry)
            if block is None:
                continue
            page_blocks.append(block)
            ocr_blocks.append(block)
        if page_blocks:
            pages.append(page_from_blocks(page_number, page_blocks, "ocr"))
    if file_type == "pdf":
        notes.append("PaddleOCR processed PDF pages for OCR fallback coverage.")
    full_text = normalize_text("\n\n".join(page["text"] for page in pages if page["text"]))
    return build_result_payload(engine="python:paddleocr", pages=pages, full_text=full_text, notes=notes, warnings=["OCR completed without yielding readable page text."] if not pages else [], extractor_chain=["python:paddleocr"], ocr_blocks=ocr_blocks, metadata={"ocrBlockCount": len(ocr_blocks)})


def build_capabilities() -> Dict[str, Any]:
    modules = {
        "pypdf": module_available("pypdf"),
        "docx": module_available("docx"),
        "openpyxl": module_available("openpyxl"),
        "paddleocr": module_available("paddleocr"),
        "docling": module_available("docling"),
    }
    versions = {
        "pypdf": module_version("pypdf"),
        "python-docx": module_version("python-docx"),
        "openpyxl": module_version("openpyxl"),
        "paddleocr": module_version("paddleocr"),
        "docling": module_version("docling"),
    }
    reasons: List[str] = []
    if not modules["pypdf"]:
        reasons.append("pypdf is unavailable.")
    if not modules["docx"]:
        reasons.append("python-docx is unavailable.")
    if not modules["openpyxl"]:
        reasons.append("openpyxl is unavailable.")
    if not modules["paddleocr"]:
        reasons.append("PaddleOCR is unavailable.")
    if not modules["docling"]:
        reasons.append("Docling is unavailable.")
    return {
        "available": any(modules.values()),
        "nativeReady": modules["pypdf"] or modules["docx"] or modules["openpyxl"],
        "ocrReady": modules["paddleocr"] or modules["docling"],
        "modules": modules,
        "versions": versions,
        "reasons": reasons,
    }


def normalize_text(value: str, *, preserve_line_breaks: bool = True) -> str:
    text = str(value or "").replace("\r", "").replace("\u00a0", " ")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    if preserve_line_breaks:
        text = re.sub(r"\n{3,}", "\n\n", text)
        return text.strip()
    return re.sub(r"\s+", " ", text).strip()


def normalize_pdf_text(value: str) -> str:
    lines: List[str] = []
    for raw_line in str(value or "").replace("\r", "").split("\n"):
        line = re.sub(r"[ \t]{2,}", " ", raw_line.replace("\u00a0", " ")).strip()
        if not line:
            if lines and lines[-1] != "":
                lines.append("")
            continue
        if lines and lines[-1].endswith("-") and line[:1].islower():
            lines[-1] = f"{lines[-1][:-1]}{line}"
            continue
        lines.append(line)
    return normalize_text("\n".join(lines))


def dedupe_strings(values: Iterable[str]) -> List[str]:
    seen = set()
    ordered: List[str] = []
    for value in values:
        normalized = normalize_text(value, preserve_line_breaks=False)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(normalized)
    return ordered


def build_language_hints(text: str) -> List[str]:
    hints: List[str] = []
    if re.search(r"[A-Za-z]", text or ""):
        hints.append("en")
    if re.search(r"[\u0600-\u06FF]", text or ""):
        hints.append("ar")
    return hints


def normalize_table_rows(rows: Iterable[Sequence[Any]]) -> List[List[str]]:
    normalized_rows: List[List[str]] = []
    max_len = 0
    for row in rows:
        converted = [normalize_text(str(cell), preserve_line_breaks=False) if cell is not None else "" for cell in row]
        while converted and converted[-1] == "":
            converted.pop()
        if not any(converted):
            continue
        max_len = max(max_len, len(converted))
        normalized_rows.append(converted)
    if max_len > 0:
        for row in normalized_rows:
            if len(row) < max_len:
                row.extend([""] * (max_len - len(row)))
    return normalized_rows


def rows_to_table_text(rows: Iterable[Sequence[Any]]) -> str:
    normalized_rows = normalize_table_rows(rows)
    return normalize_text("\n".join(" | ".join(cell for cell in row if cell != "") for row in normalized_rows))


def looks_like_heading(line: str) -> bool:
    trimmed = (line or "").strip()
    if not trimmed or len(trimmed) > 160:
        return False
    if re.match(r"^(chapter|section|unit|module|lesson|topic|abstract|introduction|conclusion)\b", trimmed, re.I):
        return True
    if HEADING_PREFIX_PATTERN.match(trimmed):
        return True
    if re.match(r"^[A-Z0-9\s\-:()]{4,}$", trimmed):
        return True
    if re.match(r"^[\u0600-\u06FF0-9\s\-:()]{4,}$", trimmed) and len(trimmed) <= 80:
        return True
    return False


def looks_like_title_candidate(line: str, index: int) -> bool:
    trimmed = (line or "").strip()
    if index > 1 or not trimmed or len(trimmed) > 120:
        return False
    return bool(
        re.match(r"^[A-Z][A-Za-z0-9\s\-:()]{4,}$", trimmed)
        or re.match(r"^[\u0600-\u06FF][\u0600-\u06FF0-9\s\-:()]{3,}$", trimmed)
    )


def looks_like_list_item(line: str) -> bool:
    return bool(LIST_MARKER_PATTERN.match((line or "").strip()))


def looks_like_table_row(line: str) -> bool:
    trimmed = (line or "").strip()
    if not trimmed:
        return False
    return "\t" in trimmed or "|" in trimmed or re.search(r"\s{3,}", trimmed) is not None


def strip_list_marker(line: str) -> str:
    return LIST_MARKER_PATTERN.sub("", line or "").strip()


def infer_heading_metadata(line: str, index: int) -> Optional[Tuple[str, int]]:
    trimmed = (line or "").strip()
    if not looks_like_heading(trimmed):
        if looks_like_title_candidate(trimmed, index):
            return ("title", 1)
        return None
    numbered_match = HEADING_PREFIX_PATTERN.match(trimmed)
    if numbered_match:
        depth = len(numbered_match.group(1).rstrip(".").split("."))
        return ("heading" if depth <= 1 else "subheading", min(6, depth + 1))
    if looks_like_title_candidate(trimmed, index):
        return ("title", 1)
    return ("heading" if re.match(r"^[A-Z0-9\s\-:()]{4,}$", trimmed) else "subheading", 2 if re.match(r"^[A-Z0-9\s\-:()]{4,}$", trimmed) else 3)


def create_block(page_number: int, order: int, text: str, source: str, *, block_type: Optional[str] = None, level: Optional[int] = None, rows: Optional[List[List[str]]] = None, confidence: Optional[float] = None, bbox: Optional[Dict[str, float]] = None, notes: Optional[List[str]] = None) -> Dict[str, Any]:
    normalized = normalize_text(text)
    if not normalized:
        normalized = text.strip()
    resolved_type = block_type or "paragraph"
    resolved_level = level
    if not block_type:
        heading = infer_heading_metadata(normalized, order - 1)
        if heading:
            resolved_type, resolved_level = heading
        elif looks_like_list_item(normalized):
            resolved_type = "list_item"
        elif looks_like_table_row(normalized):
            resolved_type = "table"
            split_rows = [re.split(r"\s{3,}|\t|\|", normalized)]
            rows = [[cell.strip() for cell in row if cell.strip()] for row in split_rows if row]
    return {
        "blockId": f"py-{source}-{page_number}-{order}",
        "type": resolved_type,
        "source": source,
        "text": normalized,
        "pageNumber": page_number,
        "order": order,
        "level": resolved_level,
        "rows": rows,
        "confidence": confidence,
        "bbox": bbox,
        "notes": notes,
    }


def block_plain_text(block: Dict[str, Any]) -> str:
    block_type = str(block.get("type") or "paragraph")
    text = normalize_text(str(block.get("text") or ""))
    if not text:
        return ""
    if block_type == "list_item":
        return f"- {strip_list_marker(text)}"
    if block_type == "table":
        rows = block.get("rows") or []
        if isinstance(rows, list) and rows:
            return "\n".join(" | ".join(str(cell).strip() for cell in row if str(cell).strip()) for row in rows)
    if block_type == "ocr_block":
        return f"OCR: {text}"
    if block_type == "note":
        return f"Note: {text}"
    return text


def lines_to_blocks(page_number: int, text: str, source: str, *, starting_order: int = 0) -> List[Dict[str, Any]]:
    lines = [line.strip() for line in normalize_text(text).split("\n") if line.strip()]
    return [create_block(page_number, starting_order + index, line, source) for index, line in enumerate(lines, start=1)]


def markdown_table_to_rows(lines: Sequence[str]) -> List[List[str]]:
    rows: List[List[str]] = []
    for line in lines:
        stripped = line.strip()
        if MARKDOWN_TABLE_SEPARATOR_PATTERN.match(stripped):
            continue
        if "|" not in stripped:
            continue
        rows.append([cell.strip() for cell in stripped.strip("|").split("|")])
    return normalize_table_rows(rows)


def markdown_to_blocks(page_number: int, markdown: str, source: str) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    lines = str(markdown or "").replace("\r", "").split("\n")
    index = 0
    order = 1
    while index < len(lines):
        line = lines[index].strip()
        index += 1
        if not line:
            continue
        table_lines = [line]
        if line.startswith("|") and "|" in line:
            while index < len(lines):
                next_line = lines[index].strip()
                if not next_line.startswith("|") or "|" not in next_line:
                    break
                table_lines.append(next_line)
                index += 1
            if len(table_lines) > 1:
                rows = markdown_table_to_rows(table_lines)
                if rows:
                    blocks.append(create_block(page_number, order, rows_to_table_text(rows), source, block_type="table", rows=rows))
                    order += 1
                    continue
        heading_match = MARKDOWN_HEADING_PATTERN.match(line)
        if heading_match:
            depth = len(heading_match.group(1))
            text = normalize_text(heading_match.group(2))
            block_type = "title" if depth == 1 else "heading" if depth <= 2 else "subheading"
            blocks.append(create_block(page_number, order, text, source, block_type=block_type, level=min(6, depth)))
            order += 1
            continue
        if line.startswith(">"):
            blocks.append(create_block(page_number, order, line.lstrip("> ").strip(), source, block_type="note"))
            order += 1
            continue
        if looks_like_list_item(line):
            blocks.append(create_block(page_number, order, line, source, block_type="list_item"))
            order += 1
            continue
        blocks.append(create_block(page_number, order, line, source))
        order += 1
    return blocks


def default_page_label(file_type: str, page_number: int, explicit: Optional[str] = None) -> str:
    if explicit:
        return explicit
    if file_type == "pdf":
        return f"Page {page_number}"
    if file_type in ("xlsx", "xls", "csv"):
        return f"Sheet {page_number}"
    if file_type == "image":
        return f"Image {page_number}"
    return f"Section {page_number}"


def page_from_blocks(page_number: int, blocks: List[Dict[str, Any]], source_kind: str, *, label: Optional[str] = None, file_type: str = "pdf") -> Dict[str, Any]:
    sorted_blocks = sorted(blocks, key=lambda item: int(item.get("order") or 0))
    page_text = normalize_text("\n".join(block_plain_text(block) for block in sorted_blocks if block_plain_text(block)))
    heading_candidates = [str(block.get("text") or "") for block in sorted_blocks if str(block.get("type") or "") in ("title", "heading", "subheading")][:10]
    return {"pageNumber": page_number, "label": default_page_label(file_type, page_number, label), "text": page_text, "headingCandidates": heading_candidates, "blocks": sorted_blocks, "sourceKind": source_kind}


def page_payload(page_number: int, text: str, source_kind: str, *, label: Optional[str] = None, file_type: str = "pdf") -> Dict[str, Any]:
    return page_from_blocks(page_number, lines_to_blocks(page_number, text, source_kind), source_kind, label=label, file_type=file_type)


def build_extraction_result(
    engine: str,
    pages: List[Dict[str, Any]],
    *,
    notes: Optional[List[str]] = None,
    warnings: Optional[List[str]] = None,
    ocr_blocks: Optional[List[Dict[str, Any]]] = None,
    extractor_chain: Optional[List[str]] = None,
    fallback_chain: Optional[List[str]] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    normalized_pages = [
        page
        for page in sorted(pages, key=lambda item: int(item.get("pageNumber") or 0))
        if normalize_text(str(page.get("text") or ""))
    ]
    full_text = normalize_text("\n\n".join(str(page.get("text") or "") for page in normalized_pages))
    payload: Dict[str, Any] = {
        "engine": engine,
        "pages": normalized_pages,
        "fullText": full_text,
        "languageHints": build_language_hints(full_text),
        "notes": dedupe_strings(notes or []),
        "warnings": dedupe_strings(warnings or []),
        "extractorChain": dedupe_strings(extractor_chain or [engine]),
        "fallbackChain": dedupe_strings(fallback_chain or []),
        "pageCount": len(normalized_pages),
        "textLength": len(full_text),
        "tablesDetected": sum(
            1
            for page in normalized_pages
            for block in page.get("blocks", [])
            if str(block.get("type") or "") == "table"
        ),
        "listsDetected": sum(
            1
            for page in normalized_pages
            for block in page.get("blocks", [])
            if str(block.get("type") or "") == "list_item"
        ),
        "metadata": metadata or {},
    }
    if ocr_blocks is not None:
        payload["ocrBlocks"] = ocr_blocks or []
        payload["ocrUsed"] = len(ocr_blocks or []) > 0
    return payload


def text_is_weak(text: str, *, min_chars: int = 80) -> bool:
    normalized = normalize_text(text, preserve_line_breaks=False)
    alnum = re.sub(r"[^0-9A-Za-z\u0600-\u06FF]", "", normalized)
    return len(alnum) < min_chars


def extract_pdf_page_text(page: Any) -> str:
    try:
        text = page.extract_text(extraction_mode="layout") or ""
    except TypeError:
        text = ""
    except Exception:
        text = ""
    if not text:
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
    return normalize_pdf_text(text)


def extract_pdf_native(source_path: str) -> Dict[str, Any]:
    from pypdf import PdfReader  # type: ignore

    reader = PdfReader(source_path)
    pages: List[Dict[str, Any]] = []
    weak_pages: List[int] = []
    for index, page in enumerate(reader.pages, start=1):
        text = extract_pdf_page_text(page)
        if text_is_weak(text):
            weak_pages.append(index)
        pages.append(page_payload(index, text, "native", label=f"Page {index}", file_type="pdf"))
    warnings: List[str] = []
    if weak_pages:
        warnings.append("Native PDF extraction produced sparse text on pages: " + ", ".join(str(page_no) for page_no in weak_pages))
    result = build_extraction_result(
        "python:pypdf",
        pages,
        warnings=warnings,
        extractor_chain=["python:pypdf:layout", "python:pypdf:plain-fallback"],
        metadata={"weakPages": weak_pages, "unitKind": "page"},
    )
    result["weakPages"] = weak_pages
    return result


def paragraph_has_numbering(paragraph: Any) -> bool:
    try:
        ppr = paragraph._p.pPr  # type: ignore[attr-defined]
        return bool(getattr(ppr, "numPr", None))
    except Exception:
        return False


def docx_paragraph_to_block(page_number: int, order: int, paragraph: Any) -> Optional[Dict[str, Any]]:
    text = normalize_text(paragraph.text or "")
    if not text:
        return None
    style_name = normalize_text(getattr(getattr(paragraph, "style", None), "name", ""), preserve_line_breaks=False).lower()
    if style_name.startswith("title"):
        return create_block(page_number, order, text, "native", block_type="title", level=1)
    if style_name.startswith("heading"):
        match = re.search(r"(\d+)", style_name)
        level = int(match.group(1)) if match else 1
        block_type = "heading" if level <= 1 else "subheading"
        return create_block(page_number, order, text, "native", block_type=block_type, level=min(6, level + 1))
    if paragraph_has_numbering(paragraph) or "list" in style_name or "bullet" in style_name:
        return create_block(page_number, order, text, "native", block_type="list_item")
    return create_block(page_number, order, text, "native")


def iter_docx_body_items(document: Any) -> Iterable[Any]:
    from docx.table import Table  # type: ignore
    from docx.text.paragraph import Paragraph  # type: ignore

    # Walk the real XML body so paragraphs and tables stay in author order.
    for child in document.element.body.iterchildren():
        if child.tag.endswith("}p"):
            yield Paragraph(child, document)
        elif child.tag.endswith("}tbl"):
            yield Table(child, document)


def extract_docx_native(source_path: str) -> Dict[str, Any]:
    from docx import Document  # type: ignore
    from docx.table import Table  # type: ignore

    document = Document(source_path)
    blocks: List[Dict[str, Any]] = []
    order = 1
    for item in iter_docx_body_items(document):
        if isinstance(item, Table):
            rows = normalize_table_rows([[cell.text for cell in row.cells] for row in item.rows])
            if rows:
                blocks.append(create_block(1, order, rows_to_table_text(rows), "native", block_type="table", rows=rows))
                order += 1
            continue
        block = docx_paragraph_to_block(1, order, item)
        if block is None:
            continue
        blocks.append(block)
        order += 1
    notes: List[str] = []
    shape_count = len(getattr(document, "inline_shapes", []) or [])
    if shape_count > 0:
        notes.append(f"DOCX contains {shape_count} inline image object(s).")
    page = page_from_blocks(1, blocks, "native", label="Section 1", file_type="docx")
    return build_extraction_result(
        "python:python-docx",
        [page],
        notes=notes,
        warnings=["DOCX extraction yielded no readable body blocks."] if not blocks else [],
        extractor_chain=["python:python-docx:body-order"],
        metadata={"unitKind": "section"},
    )


def sheet_rows_to_blocks(page_number: int, sheet_name: str, rows: List[List[str]]) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = [create_block(page_number, 1, f"Sheet {page_number}: {sheet_name}", "native", block_type="heading", level=2)]
    if not rows:
        blocks.append(create_block(page_number, 2, "This sheet has no non-empty cells.", "native", block_type="note"))
        return blocks
    max_columns = max(len(row) for row in rows)
    blocks.append(create_block(page_number, 2, f"Rows: {len(rows)} | Columns: {max_columns}", "native", block_type="metadata"))
    order = 3
    for chunk_start in range(0, len(rows), 25):
        chunk = rows[chunk_start:chunk_start + 25]
        row_start = chunk_start + 1
        row_end = chunk_start + len(chunk)
        blocks.append(create_block(page_number, order, rows_to_table_text(chunk), "native", block_type="table", rows=chunk, notes=[f"Rows {row_start}-{row_end} from {sheet_name}."]))
        order += 1
    return blocks


def extract_spreadsheet_native(source_path: str, file_type: str) -> Dict[str, Any]:
    pages: List[Dict[str, Any]] = []
    if file_type == "csv":
        with open(source_path, "r", encoding="utf-8", errors="ignore", newline="") as handle:
            rows = normalize_table_rows(csv.reader(handle))
        blocks = sheet_rows_to_blocks(1, "CSV Data", rows)
        pages.append(page_from_blocks(1, blocks, "native", label="Sheet 1: CSV Data", file_type=file_type))
        return build_extraction_result(
            "python:csv",
            pages,
            extractor_chain=["python:csv:native"],
            metadata={"unitKind": "sheet", "sheetCount": 1},
        )

    from openpyxl import load_workbook  # type: ignore

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    for index, sheet_name in enumerate(workbook.sheetnames, start=1):
        rows = normalize_table_rows(workbook[sheet_name].iter_rows(values_only=True))
        blocks = sheet_rows_to_blocks(index, sheet_name, rows)
        pages.append(page_from_blocks(index, blocks, "native", label=f"Sheet {index}: {sheet_name}", file_type=file_type))
    return build_extraction_result(
        "python:openpyxl",
        pages,
        warnings=["Spreadsheet extraction found no populated worksheets."] if not pages else [],
        extractor_chain=["python:openpyxl:worksheet-native"],
        metadata={"unitKind": "sheet", "sheetCount": len(pages)},
    )


def extract_plain_text(source_path: str) -> Dict[str, Any]:
    text = normalize_text(Path(source_path).read_text(encoding="utf-8", errors="ignore"))
    return build_extraction_result(
        "python:text",
        [page_payload(1, text, "native", label="Section 1", file_type="txt")],
        extractor_chain=["python:text:native"],
        metadata={"unitKind": "section"},
    )


def build_paddleocr_engine() -> Tuple[Any, List[str]]:
    from paddleocr import PaddleOCR  # type: ignore

    notes: List[str] = []
    last_error: Optional[str] = None
    for lang in ("ar", "en"):
        try:
            engine = PaddleOCR(lang=lang, use_doc_orientation_classify=True, use_doc_unwarping=True, use_textline_orientation=True)
            notes.append(f"PaddleOCR initialized with language profile: {lang}")
            return engine, notes
        except Exception as error:
            last_error = str(error)
    if last_error:
        notes.append(f"PaddleOCR language profile fallback triggered: {last_error}")
    notes.append("PaddleOCR fell back to default initialization.")
    return PaddleOCR(lang="en"), notes


def mapping_from_value(value: Any) -> Optional[Dict[str, Any]]:
    if isinstance(value, Mapping):
        return dict(value)
    for method_name in ("model_dump", "to_dict", "dict"):
        method = getattr(value, method_name, None)
        if callable(method):
            try:
                converted = method()
                if isinstance(converted, Mapping):
                    return dict(converted)
            except Exception:
                pass
    if hasattr(value, "__dict__"):
        try:
            return {key: item for key, item in vars(value).items() if not key.startswith("_")}
        except Exception:
            return None
    return None


def looks_like_ocr_line(entry: Any) -> bool:
    return isinstance(entry, (list, tuple)) and len(entry) >= 2 and isinstance(entry[0], (list, tuple))


def normalize_prediction_page(value: Any, default_page_number: int) -> Optional[Tuple[int, List[Any]]]:
    if isinstance(value, list) and value and all(looks_like_ocr_line(entry) for entry in value):
        return default_page_number, list(value)
    record = mapping_from_value(value)
    if record is None:
        return None
    nested = record.get("ocr_result") or record.get("result") or record.get("res")
    if isinstance(nested, list) and nested and all(looks_like_ocr_line(entry) for entry in nested):
        page_number = int(record.get("page_number") or record.get("page_no") or record.get("page_id") or record.get("page_index") or default_page_number)
        return page_number, list(nested)
    texts = record.get("rec_texts") or record.get("texts") or record.get("text") or record.get("ocr_texts")
    boxes = record.get("dt_polys") or record.get("polys") or record.get("boxes") or record.get("dt_boxes")
    scores = record.get("rec_scores") or record.get("scores")
    if isinstance(texts, str):
        texts = [texts]
    if not isinstance(texts, list):
        return None
    page_number = int(record.get("page_number") or record.get("page_no") or record.get("page_id") or record.get("page_index") or default_page_number)
    entries: List[Any] = []
    for index, text in enumerate(texts):
        bbox = boxes[index] if isinstance(boxes, list) and index < len(boxes) else []
        score = scores[index] if isinstance(scores, list) and index < len(scores) else None
        entries.append([bbox, [text, score]])
    return page_number, entries


def iter_ocr_pages(result: Any) -> List[Tuple[int, List[Any]]]:
    if isinstance(result, list):
        if result and all(looks_like_ocr_line(entry) for entry in result):
            return [(1, list(result))]
        pages: List[Tuple[int, List[Any]]] = []
        for index, page in enumerate(result, start=1):
            normalized = normalize_prediction_page(page, index)
            if normalized is not None:
                pages.append(normalized)
                continue
            if isinstance(page, list) and page and all(looks_like_ocr_line(entry) for entry in page):
                pages.append((index, page))
        return pages
    normalized = normalize_prediction_page(result, 1)
    return [normalized] if normalized is not None else []


def bbox_from_points(points: Sequence[Sequence[float]]) -> Optional[Dict[str, float]]:
    if not points:
        return None
    x_values = [float(point[0]) for point in points if len(point) >= 2]
    y_values = [float(point[1]) for point in points if len(point) >= 2]
    if not x_values or not y_values:
        return None
    return {"x": min(x_values), "y": min(y_values), "width": max(x_values) - min(x_values), "height": max(y_values) - min(y_values)}


def ocr_entry_to_block(page_number: int, order: int, entry: Sequence[Any]) -> Optional[Dict[str, Any]]:
    if len(entry) < 2:
        return None
    bbox = entry[0] if isinstance(entry[0], (list, tuple)) else []
    payload = entry[1] if isinstance(entry[1], (list, tuple)) else [entry[1], None]
    text = normalize_text(str(payload[0] if len(payload) > 0 else ""))
    if not text:
        return None
    try:
        confidence = float(payload[1]) if len(payload) > 1 and payload[1] is not None else None
    except Exception:
        confidence = None
    heading = infer_heading_metadata(text, order - 1)
    block_type = "ocr_block"
    level = None
    if heading:
        block_type, level = heading
    elif looks_like_list_item(text):
        block_type = "list_item"
    elif looks_like_table_row(text):
        block_type = "table"
    rows = [re.split(r"\s{3,}|\t|\|", text)] if block_type == "table" else None
    if rows is not None:
        rows = [[cell.strip() for cell in row if cell.strip()] for row in rows if row]
    return create_block(page_number, order, text, "ocr", block_type=block_type, level=level, rows=rows, confidence=confidence, bbox=bbox_from_points(bbox))


def extract_ocr_document(source_path: str, file_type: str) -> Dict[str, Any]:
    engine, notes = build_paddleocr_engine()
    try:
        if hasattr(engine, "predict"):
            result = engine.predict(source_path, use_doc_orientation_classify=True, use_doc_unwarping=True, use_textline_orientation=True)
        else:
            result = engine.ocr(source_path, cls=True)
    finally:
        close_method = getattr(engine, "close", None)
        if callable(close_method):
            try:
                close_method()
            except Exception:
                pass
    pages: List[Dict[str, Any]] = []
    ocr_blocks: List[Dict[str, Any]] = []
    for default_index, (page_number, page_entries) in enumerate(iter_ocr_pages(result), start=1):
        effective_page_number = page_number or default_index
        blocks: List[Dict[str, Any]] = []
        for entry_index, entry in enumerate(page_entries, start=1):
            block = ocr_entry_to_block(effective_page_number, entry_index, entry)
            if block is None:
                continue
            blocks.append(block)
            ocr_blocks.append(block)
        if blocks:
            pages.append(page_from_blocks(effective_page_number, blocks, "ocr", label=default_page_label(file_type, effective_page_number), file_type=file_type))
    if file_type == "pdf":
        notes.append("PaddleOCR was used as a page-level fallback for PDF extraction.")
    elif file_type == "image":
        notes.append("PaddleOCR was used as the primary extractor for the uploaded image.")
    return build_extraction_result(
        "python:paddleocr",
        pages,
        notes=notes,
        warnings=["OCR completed without yielding readable blocks."] if not pages else [],
        ocr_blocks=ocr_blocks,
        extractor_chain=["python:paddleocr"],
        metadata={"ocrBlockCount": len(ocr_blocks), "unitKind": "page"},
    )


def extract_docling_page_fragments(value: Any, fragments: Dict[int, List[str]]) -> None:
    record = mapping_from_value(value)
    if record is not None:
        page_number_raw = record.get("page_number") or record.get("page_no") or record.get("page") or record.get("page_id")
        try:
            page_number = int(page_number_raw) if page_number_raw is not None else None
        except Exception:
            page_number = None
        if page_number is not None:
            for key in ("text", "content", "markdown", "body"):
                candidate = record.get(key)
                if isinstance(candidate, str):
                    normalized = normalize_text(candidate)
                    if normalized:
                        fragments[page_number].append(normalized)
        for child in record.values():
            extract_docling_page_fragments(child, fragments)
        return
    if isinstance(value, list):
        for item in value:
            extract_docling_page_fragments(item, fragments)


def build_docling_raw_pages(file_type: str, markdown: Optional[str], text: Optional[str], structured: Any) -> List[Dict[str, Any]]:
    fragments: Dict[int, List[str]] = defaultdict(list)
    extract_docling_page_fragments(structured, fragments)
    if fragments:
        pages: List[Dict[str, Any]] = []
        for page_number in sorted(fragments):
            joined = normalize_text("\n\n".join(dedupe_strings(fragments[page_number])))
            if not joined:
                continue
            pages.append({"pageNumber": page_number, "label": default_page_label(file_type, page_number), "text": joined})
        return pages
    fallback_text = normalize_text(markdown or text or "")
    if not fallback_text:
        return []
    return [{"pageNumber": 1, "label": default_page_label(file_type, 1), "text": fallback_text}]


def load_docling_payload(source_path: str, file_type: str) -> Optional[Dict[str, Any]]:
    try:
        from docling.document_converter import DocumentConverter  # type: ignore
    except Exception:
        return None
    try:
        conversion = DocumentConverter().convert(source_path)
        document = getattr(conversion, "document", None) or getattr(conversion, "legacy_document", None)
        if document is None:
            return None
        markdown = document.export_to_markdown() if hasattr(document, "export_to_markdown") else None
        if hasattr(document, "export_to_dict"):
            structured = document.export_to_dict()
        elif hasattr(document, "export_to_json"):
            try:
                structured = json.loads(document.export_to_json())
            except Exception:
                structured = None
        else:
            structured = None
        text = document.export_to_text() if hasattr(document, "export_to_text") else None
        normalized_text = normalize_text(text or markdown or "")
        pages = build_docling_raw_pages(file_type, markdown, text, structured)
        return {
            "engine": "python:docling",
            "markdown": markdown,
            "structured": structured,
            "text": normalized_text,
            "pages": pages,
            "notes": ["Docling structural conversion completed."],
            "warnings": [],
            "extractorChain": ["python:docling:layout-aware"],
            "pageCount": len(pages),
            "textLength": len(normalized_text),
        }
    except Exception as error:
        return {
            "engine": "python:docling",
            "markdown": None,
            "structured": None,
            "text": None,
            "pages": [],
            "notes": [],
            "warnings": [f"Docling conversion failed: {error}"],
            "extractorChain": ["python:docling:layout-aware"],
            "pageCount": 0,
            "textLength": 0,
            "error": str(error),
        }


def docling_pages_to_payload_pages(docling_payload: Dict[str, Any], source_kind: str, file_type: str) -> List[Dict[str, Any]]:
    raw_pages = docling_payload.get("pages") or []
    if not isinstance(raw_pages, list):
        return []
    pages: List[Dict[str, Any]] = []
    for raw_page in raw_pages:
        page_number = int(raw_page.get("pageNumber") or 1)
        text = normalize_text(str(raw_page.get("text") or ""))
        if not text:
            continue
        label = normalize_text(str(raw_page.get("label") or ""), preserve_line_breaks=False) or default_page_label(file_type, page_number)
        blocks = markdown_to_blocks(page_number, text, source_kind) if "#" in text or "|" in text else lines_to_blocks(page_number, text, source_kind)
        pages.append(page_from_blocks(page_number, blocks, source_kind, label=label, file_type=file_type))
    return pages


def merge_native_with_docling(native: Optional[Dict[str, Any]], docling_payload: Optional[Dict[str, Any]], file_type: str) -> Optional[Dict[str, Any]]:
    if not docling_payload:
        return native
    docling_pages = docling_pages_to_payload_pages(docling_payload, "native", file_type)
    if not docling_pages:
        return native
    if native is None or not native.get("pages"):
        return build_extraction_result(
            "python:docling-native-fallback",
            docling_pages,
            notes=["Docling supplied the primary structured text output."],
            extractor_chain=["python:docling:layout-aware"],
            fallback_chain=["docling:primary"],
            metadata={"doclingPrimary": True},
        )
    native_pages = {int(page.get("pageNumber") or 0): page for page in native.get("pages", [])}
    supplemented_pages: List[int] = []
    # Docling supplements only pages that are missing or materially weak so the
    # format-native parser remains the primary owner when it already extracted
    # good text.
    for docling_page in docling_pages:
        page_number = int(docling_page.get("pageNumber") or 0)
        existing = native_pages.get(page_number)
        if existing is None:
            native_pages[page_number] = docling_page
            supplemented_pages.append(page_number)
            continue
        native_text = normalize_text(str(existing.get("text") or ""))
        docling_text = normalize_text(str(docling_page.get("text") or ""))
        if text_is_weak(native_text) and len(docling_text) > len(native_text) + 40:
            native_pages[page_number] = {**docling_page, "label": existing.get("label") or docling_page.get("label")}
            supplemented_pages.append(page_number)
    notes = list(native.get("notes") or [])
    warnings = list(native.get("warnings") or [])
    if supplemented_pages:
        notes.append("Docling supplemented weak native pages: " + ", ".join(str(page_no) for page_no in sorted(set(supplemented_pages))))
    metadata = dict(native.get("metadata") or {})
    metadata["doclingSupplementedPages"] = sorted(set(supplemented_pages))
    return build_extraction_result(
        f"{native.get('engine', 'python:native')}+docling",
        list(native_pages.values()),
        notes=notes,
        warnings=warnings,
        extractor_chain=[*(native.get("extractorChain") or [native.get("engine", "python:native")]), "python:docling:layout-aware"],
        fallback_chain=[*(native.get("fallbackChain") or []), "docling:weak-page-supplement"],
        metadata=metadata,
    )


def merge_ocr_with_docling(ocr: Optional[Dict[str, Any]], docling_payload: Optional[Dict[str, Any]], file_type: str) -> Optional[Dict[str, Any]]:
    if not docling_payload:
        return ocr
    docling_pages = docling_pages_to_payload_pages(docling_payload, "ocr", file_type)
    if not docling_pages:
        return ocr
    if ocr is None or not ocr.get("pages"):
        return build_extraction_result(
            "python:docling-ocr-fallback",
            docling_pages,
            notes=["Docling provided OCR/layout-aware fallback text."],
            ocr_blocks=[],
            extractor_chain=["python:docling:layout-aware"],
            fallback_chain=["docling:ocr-fallback"],
            metadata={"doclingPrimary": True},
        )
    by_page = {int(page.get("pageNumber") or 0): page for page in ocr.get("pages", [])}
    supplemented_pages: List[int] = []
    for docling_page in docling_pages:
        page_number = int(docling_page.get("pageNumber") or 0)
        if page_number not in by_page and normalize_text(str(docling_page.get("text") or "")):
            by_page[page_number] = docling_page
            supplemented_pages.append(page_number)
    notes = list(ocr.get("notes") or [])
    warnings = list(ocr.get("warnings") or [])
    if supplemented_pages:
        notes.append("Docling filled OCR gaps on pages: " + ", ".join(str(page_no) for page_no in sorted(set(supplemented_pages))))
    metadata = dict(ocr.get("metadata") or {})
    metadata["doclingSupplementedPages"] = sorted(set(supplemented_pages))
    return build_extraction_result(
        f"{ocr.get('engine', 'python:ocr')}+docling",
        list(by_page.values()),
        notes=notes,
        warnings=warnings,
        ocr_blocks=list(ocr.get("ocrBlocks") or []),
        extractor_chain=[*(ocr.get("extractorChain") or [ocr.get("engine", "python:ocr")]), "python:docling:layout-aware"],
        fallback_chain=[*(ocr.get("fallbackChain") or []), "docling:ocr-gap-fill"],
        metadata=metadata,
    )


def should_load_docling(file_type: str, mode: str) -> bool:
    if file_type in ("pdf", "docx", "image"):
        return True
    return mode == "ocr"


def extract_document(payload: Dict[str, Any]) -> Dict[str, Any]:
    source_path = str(payload.get("sourcePath") or "")
    file_type = str(payload.get("fileType") or "").lower()
    mode = str(payload.get("mode") or "native")
    capabilities = build_capabilities()
    notes: List[str] = []
    errors: List[str] = []
    attempted_layers: List[str] = []
    native: Optional[Dict[str, Any]] = None
    ocr: Optional[Dict[str, Any]] = None
    docling_payload: Optional[Dict[str, Any]] = None

    if capabilities["modules"].get("docling") and should_load_docling(file_type, mode):
        attempted_layers.append("docling:preload")
        docling_payload = load_docling_payload(source_path, file_type)
        if docling_payload:
            notes.extend(docling_payload.get("notes") or [])
            notes.extend(docling_payload.get("warnings") or [])
            if docling_payload.get("error"):
                errors.append(f"docling-load-failed: {docling_payload['error']}")

    try:
        if mode in ("native", "hybrid"):
            if file_type == "pdf" and capabilities["modules"].get("pypdf"):
                attempted_layers.append("native:pypdf")
                native = extract_pdf_native(source_path)
            elif file_type == "docx" and capabilities["modules"].get("docx"):
                attempted_layers.append("native:python-docx")
                native = extract_docx_native(source_path)
            elif file_type == "xlsx" and capabilities["modules"].get("openpyxl"):
                attempted_layers.append("native:openpyxl")
                native = extract_spreadsheet_native(source_path, file_type)
            elif file_type == "csv":
                attempted_layers.append("native:csv")
                native = extract_spreadsheet_native(source_path, file_type)
            elif file_type == "txt":
                attempted_layers.append("native:text")
                native = extract_plain_text(source_path)
            else:
                attempted_layers.append("native:skipped")
            native = merge_native_with_docling(native, docling_payload, file_type)
        else:
            attempted_layers.append("native:skipped")
    except Exception as error:
        errors.append(f"native-extract-failed: {error}")

    try:
        if mode in ("ocr", "hybrid"):
            if file_type in ("image", "pdf") and capabilities["modules"].get("paddleocr"):
                attempted_layers.append("ocr:paddleocr")
                ocr = extract_ocr_document(source_path, file_type)
            else:
                attempted_layers.append("ocr:skipped")
            ocr = merge_ocr_with_docling(ocr, docling_payload, file_type)
        else:
            attempted_layers.append("ocr:skipped")
    except Exception as error:
        errors.append(f"ocr-extract-failed: {error}")
        ocr = merge_ocr_with_docling(ocr, docling_payload, file_type)

    combined_notes = dedupe_strings([
        *notes,
        *((native.get("notes") or []) if native else []),
        *((ocr.get("notes") or []) if ocr else []),
        f"Attempted layers: {', '.join(attempted_layers)}",
    ])
    combined_warnings = dedupe_strings([
        *((native.get("warnings") or []) if native else []),
        *((ocr.get("warnings") or []) if ocr else []),
    ])
    return {
        "ok": native is not None or ocr is not None or docling_payload is not None,
        "native": native,
        "ocr": ocr,
        "docling": docling_payload,
        "notes": combined_notes,
        "warnings": combined_warnings,
        "errors": dedupe_strings(errors),
        "capabilities": capabilities,
        "attemptedLayers": attempted_layers,
        "fileType": file_type,
        "mode": mode,
    }


def main() -> int:
    if len(sys.argv) != 4:
        return 2
    command = sys.argv[1]
    input_path = sys.argv[2]
    output_path = sys.argv[3]
    if command == "detect":
        write_json(output_path, build_capabilities())
        return 0
    if command == "extract":
        write_json(output_path, extract_document(load_json(input_path)))
        return 0
    write_json(output_path, {"error": f"Unknown command: {command}"})
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
